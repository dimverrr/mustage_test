import requests
from bs4 import BeautifulSoup
from sqlalchemy.orm import Session
from classes.db import Engine, ExchangeRate
import openpyxl
import aioschedule
import asyncio
from datetime import date
from sqlalchemy import func
import os
from classes.logger import Logger
from config_reader import config
from openpyxl.styles import PatternFill

logger = Logger()

def get_today_rate_from_db(today: date):
    """
    Retrieves exchange rate that was collected today
    Args:
        date (date): todays date
    Returns:
        list: todays exchange rates
    """
    with Session(autoflush=False, bind=Engine().engine) as db:
        result = (
            db.query(ExchangeRate.date, ExchangeRate.rate)
            .filter(func.date(ExchangeRate.date) == today)
            .all()
        )

    if not result:
        logger.error("Impossible to retrieve data from db")

    return result


def write_to_file():
    """
    Writes exchange rate to the file with todays date 

    Returns:
        str: name of file, where exchange rates are saved
    """
    today = date.today()

    today_rates = get_today_rate_from_db(today)

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    headers = ["datetime", "exchange_rate"]
    sheet.append(headers)

    sheet.column_dimensions["A"].width = 20
    sheet["A1"].fill = PatternFill("solid", start_color="FFFF00")
    sheet.column_dimensions["B"].width = 20
    sheet["B1"].fill = PatternFill("solid", start_color="FFFF00")

    for row in today_rates:
        sheet.append(list(row))

    file_name = f"Exchange USD-UAH {today}.xlsx"
    path = os.path.join("rates", file_name)
    workbook.save(path)

    return path


def save_exchange_rate(rate: float):
    """
    Saves new exchange rate value into database

    Args:
        rate (float): exchange rate of currency
    """
    with Session(autoflush=False, bind=Engine().engine) as db:
        exchange_rate = ExchangeRate(rate=rate)
        db.add(exchange_rate)
        db.commit()

async def exchange_request_uah():
    """
    Retrieves data from currency exchange website

    Returns:
        none: returns None in case of no data 
    """
    url = config.url.get_secret_value()

    response = requests.get(url)
    response.raise_for_status()

    soup = BeautifulSoup(response.text, "html.parser")

    last_price_element = soup.select_one("[data-last-price]")

    if last_price_element:
        last_price = last_price_element.get("data-last-price")
        logger.info(f"Last rate is {last_price}")
        save_exchange_rate(float(last_price))
    else:
        logger.error("Unable to retrieve exchange rate from site")
        return None

async def schedule_exchange_retrieve_request():
    """
    Schedules time when request for new exchange rate will be perfomed
    """
    aioschedule.every().hour.at(":00").do(exchange_request_uah)
    while True:
        await aioschedule.run_pending()
        await asyncio.sleep(1)


